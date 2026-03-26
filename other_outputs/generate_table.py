import os
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

BASE = r"C:\Users\swl00\IFPRI Dropbox\Weilun Shi\Google fund\Analysis\2.source_code\Step5_Geo_RF_trial\Food_Crisis_Cluster"

def read_csv(path):
    with open(path, newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))

def avg(vals):
    vals = [v for v in vals if v is not None]
    return sum(vals) / len(vals) if vals else None

def collect_model_metrics(rows):
    sp, sr, sf = [], [], []
    pp, pr, pf = [], [], []
    for r in rows:
        sp.append(float(r["precision(1)"]))
        sr.append(float(r["recall(1)"]))
        sf.append(float(r["f1(1)"]))
        pp.append(float(r["precision_base(1)"]))
        pr.append(float(r["recall_base(1)"]))
        pf.append(float(r["f1_base(1)"]))
    return {
        "split_precision": avg(sp), "split_recall": avg(sr), "split_f1": avg(sf),
        "pool_precision": avg(pp), "pool_recall": avg(pr), "pool_f1": avg(pf),
        "n": len(rows),
    }

def collect_fewsnet_metrics(rows):
    p, r, f = [], [], []
    for row in rows:
        p.append(float(row["precision(1)"]))
        r.append(float(row["recall(1)"]))
        f.append(float(row["f1(1)"]))
    return {
        "split_precision": avg(p), "split_recall": avg(r), "split_f1": avg(f),
        "pool_precision": None, "pool_recall": None, "pool_f1": None,
        "n": len(rows),
    }

def load_model_data(results_dir, prefix):
    out = {}
    for fs in (1, 2, 3):
        all_rows = []
        matched_files = []
        for fname in sorted(os.listdir(results_dir)):
            if fname.startswith(f"{prefix}fs{fs}_") and fname.endswith(".csv"):
                fpath = os.path.join(results_dir, fname)
                rows = read_csv(fpath)
                all_rows.extend(rows)
                matched_files.append((fname, len(rows)))
        if all_rows:
            out[fs] = collect_model_metrics(all_rows)
            out[fs]["files"] = matched_files
    return out

georf = load_model_data(os.path.join(BASE, "GeoRFExperiment", "GeoRFResults"), "results_df_gp_")
geoxgb = load_model_data(os.path.join(BASE, "GeoXGBExperiment", "GeoXgboostResults"), "results_df_xgb_gp_")
geodt = load_model_data(os.path.join(BASE, "GeoDTExperiment", "GeoDTResults"), "results_df_dt_gp_")

fewsnet = {}
for fs in (1, 2):
    fpath = os.path.join(BASE, "deliverables", "fewsnet_baseline_results", f"fewsnet_baseline_results_fs{fs}.csv")
    if os.path.exists(fpath):
        rows = read_csv(fpath)
        fewsnet[fs] = collect_fewsnet_metrics(rows)
        fewsnet[fs]["files"] = [(os.path.basename(fpath), len(rows))]

FS_TO_LAG = {1: 4, 2: 8, 3: 12}

MODEL_ROWS = [
    ("GeoRF", georf),
    ("GeoXGB", geoxgb),
    ("GeoDT", geodt),
    ("FEWSNET (baseline)", fewsnet),
]

print("=" * 100)
print("DATA SOURCE AUDIT")
print("=" * 100)
for model_name, data in MODEL_ROWS:
    for fs in (1, 2, 3):
        m = data.get(fs)
        if m:
            print(f"\n{model_name} fs{fs} (lag={FS_TO_LAG[fs]}) -- {m['n']} total rows from:")
            for fname, nrows in m.get("files", []):
                print(f"    {fname} ({nrows} rows)")
            print(f"  -> split:  P={m['split_precision']:.10f}  R={m['split_recall']:.10f}  F1={m['split_f1']:.10f}")
            if m['pool_precision'] is not None:
                print(f"  -> pooled: P={m['pool_precision']:.10f}  R={m['pool_recall']:.10f}  F1={m['pool_f1']:.10f}")
            else:
                print(f"  -> pooled: (none)")
        else:
            print(f"\n{model_name} fs{fs} (lag={FS_TO_LAG[fs]}) -- NO DATA")

header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
model_font = Font(bold=True, size=11)
thin = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
center_align = Alignment(horizontal="center")

def build_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws.merge_cells("C1:E1")
    ws["C1"] = "Split Model"
    ws["C1"].font = header_font
    ws["C1"].alignment = center_align
    ws["C1"].fill = header_fill

    ws.merge_cells("F1:H1")
    ws["F1"] = "Pooled Model(Non-split)"
    ws["F1"].font = header_font
    ws["F1"].alignment = center_align
    ws["F1"].fill = header_fill

    headers = ["", "lag(months)", "Precision", "Recall", "F1",
               "precision", "recall", "F1", "F1 Improvement Percentage"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin

    row = 3
    for model_name, data in MODEL_ROWS:
        first = True
        for fs in (1, 2, 3):
            lag = FS_TO_LAG[fs]
            m = data.get(fs)

            if first:
                ws.cell(row=row, column=1, value=model_name).font = model_font
            ws.cell(row=row, column=2, value=lag).alignment = center_align

            if m:
                for ci, key in enumerate(["split_precision", "split_recall", "split_f1"], 3):
                    v = m.get(key)
                    if v is not None:
                        ws.cell(row=row, column=ci, value=v)
                for ci, key in enumerate(["pool_precision", "pool_recall", "pool_f1"], 6):
                    v = m.get(key)
                    if v is not None:
                        ws.cell(row=row, column=ci, value=v)
                sf, pf = m.get("split_f1"), m.get("pool_f1")
                if sf is not None and pf is not None and pf != 0:
                    ws.cell(row=row, column=9, value=(sf - pf) / pf)
                    ws.cell(row=row, column=9).number_format = "0.00%"

            for c in range(1, 10):
                ws.cell(row=row, column=c).border = thin

            first = False
            row += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    for col in "CDEFGH":
        ws.column_dimensions[col].width = 14
    ws.column_dimensions["I"].width = 24

    return wb

wb = build_workbook()
path1 = os.path.join(BASE, "other_outputs", "Table_Format.xlsx")
wb.save(path1)
print(f"\nSaved: {path1}")

wb2 = build_workbook()
path2 = os.path.join(BASE, "other_outputs", "Model_Comparison_Table.xlsx")
wb2.save(path2)
print(f"Saved: {path2}")

print("\n" + "=" * 100)
print("FINAL CELL VALUES (full precision, written to xlsx)")
print("=" * 100)
fmt = f"{'Model':<22} {'Lag':>3} | {'Split P':>14} {'Split R':>14} {'Split F1':>14} | {'Pool P':>14} {'Pool R':>14} {'Pool F1':>14} | {'Impr%':>10}"
print(fmt)
print("-" * 130)
for model_name, data in MODEL_ROWS:
    for fs in (1, 2, 3):
        m = data.get(fs)
        label = model_name if fs == 1 else ""
        lag = FS_TO_LAG[fs]
        if m:
            sp = f"{m['split_precision']:.10f}"
            sr = f"{m['split_recall']:.10f}"
            sf = f"{m['split_f1']:.10f}"
            pp = f"{m['pool_precision']:.10f}" if m['pool_precision'] is not None else "         -"
            pr = f"{m['pool_recall']:.10f}" if m['pool_recall'] is not None else "         -"
            pf_val = m['pool_f1']
            pf = f"{pf_val:.10f}" if pf_val is not None else "         -"
            if m['split_f1'] is not None and pf_val is not None and pf_val != 0:
                imp = f"{(m['split_f1'] - pf_val) / pf_val * 100:+.6f}%"
            else:
                imp = "      -"
            print(f"{label:<22} {lag:>3} | {sp:>14} {sr:>14} {sf:>14} | {pp:>14} {pr:>14} {pf:>14} | {imp:>10}")
        else:
            print(f"{label:<22} {lag:>3} | {'        -':>14} {'        -':>14} {'        -':>14} | {'        -':>14} {'        -':>14} {'        -':>14} | {'     -':>10}")
    print("-" * 130)
