import os
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

def _candidates(suffix, fs):
    """Return list of candidate paths in priority order: new _fsN dir, then legacy dir without suffix."""
    return [
        os.path.join(BASE, f"result_partition_k40_compare_{suffix}_fs{fs}", "metrics_monthly.csv"),
        os.path.join(BASE, f"result_partition_k40_compare_{suffix}", "metrics_monthly.csv"),
    ]

RESULT_CANDIDATES = {
    ("GeoRF", 1): _candidates("GF", 1),
    ("GeoRF", 2): _candidates("GF", 2),
    ("GeoRF", 3): _candidates("GF", 3),
    ("GeoXGB", 1): _candidates("XGB", 1),
    ("GeoXGB", 2): _candidates("XGB", 2),
    ("GeoXGB", 3): _candidates("XGB", 3),
    ("GeoDT", 1): _candidates("DT", 1),
    ("GeoDT", 2): _candidates("DT", 2),
    ("GeoDT", 3): _candidates("DT", 3),
}

def resolve_path(candidates):
    for p in candidates:
        if os.path.exists(p):
            return p
    return None

FEWSNET_MAP = {
    1: os.path.join(BASE, "fewsnet_baseline_results", "fewsnet_baseline_results_fs1.csv"),
    2: os.path.join(BASE, "fewsnet_baseline_results", "fewsnet_baseline_results_fs2.csv"),
}

FS_TO_LAG = {1: 4, 2: 8, 3: 12}


def read_metrics(path):
    with open(path, newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def avg(vals):
    return sum(vals) / len(vals) if vals else None


def std(vals):
    if len(vals) < 2:
        return 0.0
    m = sum(vals) / len(vals)
    return (sum((v - m) ** 2 for v in vals) / (len(vals) - 1)) ** 0.5


def aggregate_model_metrics(path):
    rows = read_metrics(path)
    pooled = [r for r in rows if r["model"] == "pooled"]
    partitioned = [r for r in rows if r["model"] == "partitioned"]

    def extract(subset, col):
        return [float(r[col]) for r in subset]

    return {
        "split_precision": avg(extract(partitioned, "precision")),
        "split_recall": avg(extract(partitioned, "recall")),
        "split_f1": avg(extract(partitioned, "f1")),
        "split_f1_std": std(extract(partitioned, "f1")),
        "pool_precision": avg(extract(pooled, "precision")),
        "pool_recall": avg(extract(pooled, "recall")),
        "pool_f1": avg(extract(pooled, "f1")),
        "pool_f1_std": std(extract(pooled, "f1")),
        "n_months": len(partitioned),
    }


def aggregate_fewsnet(path, min_year=2021):
    rows = [r for r in read_metrics(path) if int(r["year"]) >= min_year]
    p = [float(r["precision(1)"]) for r in rows]
    r = [float(r["recall(1)"]) for r in rows]
    f = [float(r["f1(1)"]) for r in rows]
    return {
        "split_precision": avg(p),
        "split_recall": avg(r),
        "split_f1": avg(f),
        "split_f1_std": std(f),
        "pool_precision": None,
        "pool_recall": None,
        "pool_f1": None,
        "pool_f1_std": None,
        "n_months": len(rows),
    }


data = {}

print("Loading results...")
for (model, fs), candidates in RESULT_CANDIDATES.items():
    path = resolve_path(candidates)
    if path:
        data[(model, fs)] = aggregate_model_metrics(path)
        m = data[(model, fs)]
        print(f"  {model} fs{fs}: {m['n_months']} months, split_f1={m['split_f1']:.4f}, pool_f1={m['pool_f1']:.4f}  <-- {os.path.dirname(path)}")
    else:
        print(f"  {model} fs{fs}: NOT FOUND")

for fs, path in FEWSNET_MAP.items():
    if os.path.exists(path):
        data[("FEWSNET (baseline)", fs)] = aggregate_fewsnet(path)
        m = data[("FEWSNET (baseline)", fs)]
        print(f"  FEWSNET fs{fs}: {m['n_months']} quarters, f1={m['split_f1']:.4f}")

if ("FEWSNET (baseline)", 3) not in data and ("FEWSNET (baseline)", 2) in data:
    data[("FEWSNET (baseline)", 3)] = data[("FEWSNET (baseline)", 2)].copy()
    print(f"  FEWSNET fs3: extended from fs2 (8-month predictions used as 12-month proxy)")

MODELS = ["GeoRF", "GeoXGB", "GeoDT", "FEWSNET (baseline)"]

hfont = Font(bold=True, size=11)
hfill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
mfont = Font(bold=True, size=11)
thin = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
center = Alignment(horizontal="center")

wb = Workbook()
ws = wb.active
ws.title = "Sheet1"

ws.merge_cells("C1:E1")
ws["C1"] = "Split Model"
ws["C1"].font = hfont
ws["C1"].alignment = center
ws["C1"].fill = hfill

ws.merge_cells("F1:H1")
ws["F1"] = "Pooled Model(Non-split)"
ws["F1"].font = hfont
ws["F1"].alignment = center
ws["F1"].fill = hfill

headers = ["", "lag(months)", "Precision", "Recall", "F1",
           "precision", "recall", "F1", "F1 Improvement Percentage"]
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=2, column=c, value=h)
    cell.font = hfont
    cell.fill = hfill
    cell.alignment = center
    cell.border = thin

row = 3
for model in MODELS:
    first = True
    for fs in (1, 2, 3):
        lag = FS_TO_LAG[fs]
        m = data.get((model, fs))

        if first:
            ws.cell(row=row, column=1, value=model).font = mfont
        ws.cell(row=row, column=2, value=lag).alignment = center

        if m:
            for ci, key in enumerate(["split_precision", "split_recall", "split_f1"], 3):
                v = m.get(key)
                if v is not None:
                    ws.cell(row=row, column=ci, value=v)
            for ci, key in enumerate(["pool_precision", "pool_recall", "pool_f1"], 6):
                v = m.get(key)
                if v is not None:
                    ws.cell(row=row, column=ci, value=v)
            sf, pf = m.get("split_f1"), m.get("pool_f1")
            if sf is not None and pf is not None and pf != 0:
                ws.cell(row=row, column=9, value=(sf - pf) / pf)
                ws.cell(row=row, column=9).number_format = "0.00%"

        for c in range(1, 10):
            ws.cell(row=row, column=c).border = thin

        first = False
        row += 1

ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 14
for col in "CDEFGH":
    ws.column_dimensions[col].width = 14
ws.column_dimensions["I"].width = 24

out_path = os.path.join(BASE, "other_outputs", "Table_Format.xlsx")
wb.save(out_path)
print(f"\nSaved: {out_path}")

print("\n" + "=" * 130)
fmt = f"{'Model':<22} {'Lag':>3} | {'P(split)':>10} {'R(split)':>10} {'F1(split)':>12} | {'P(pool)':>10} {'R(pool)':>10} {'F1(pool)':>12} | {'Impr%':>8}"
print(fmt)
print("-" * 130)
for model in MODELS:
    for fs in (1, 2, 3):
        m = data.get((model, fs))
        label = model if fs == 1 else ""
        lag = FS_TO_LAG[fs]
        if m:
            sp = f"{m['split_precision']:.4f}" if m['split_precision'] is not None else "    -"
            sr = f"{m['split_recall']:.4f}" if m['split_recall'] is not None else "    -"
            sf = f"{m['split_f1']:.4f}+-{m['split_f1_std']:.4f}" if m['split_f1'] is not None else "    -"
            pp = f"{m['pool_precision']:.4f}" if m['pool_precision'] is not None else "    -"
            pr = f"{m['pool_recall']:.4f}" if m['pool_recall'] is not None else "    -"
            pf_val = m['pool_f1']
            pf = f"{pf_val:.4f}+-{m['pool_f1_std']:.4f}" if pf_val is not None else "    -"
            if m['split_f1'] is not None and pf_val is not None and pf_val != 0:
                imp = f"{(m['split_f1'] - pf_val) / pf_val * 100:+.2f}%"
            else:
                imp = "   -"
            print(f"{label:<22} {lag:>3} | {sp:>10} {sr:>10} {sf:>12} | {pp:>10} {pr:>10} {pf:>12} | {imp:>8}")
        else:
            print(f"{label:<22} {lag:>3} | {'   -':>10} {'   -':>10} {'       -':>12} | {'   -':>10} {'   -':>10} {'       -':>12} | {'  -':>8}")
    print("-" * 130)
