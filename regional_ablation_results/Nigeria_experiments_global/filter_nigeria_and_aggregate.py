"""
Filter monthly prediction files to Nigeria admin codes (FEWSNET_admin_code from
Nigeria.shp) and aggregate precision / recall / F1 for pooled and partitioned
models, mirroring the logic of other_outputs/aggregate_results.py.

Inputs
------
- Nigeria shapefile:
  C:\\Users\\swl00\\IFPRI Dropbox\\Weilun Shi\\Google fund\\Analysis\\1.Source Data\\
    Outcome\\FEWSNET_IPC\\FEWS NET Admin Boundaries\\Nigeria.shp
  Uses the `admin_code` column (a.k.a. FEWSNET_admin_code).

- Monthly predictions for each (model, forecasting_scope):
    main_ablation_results/march2026_main_backup_month_ind_cont3/
        result_partition_k40_compare_{GF|XGB|DT}_fs{1|2|3}/predictions_monthly.csv

  Columns: FEWSNET_admin_code, month_start, partition_id, y_true,
           y_pred_pooled, y_pred_partitioned

Outputs (written to Nigeria_experiments/)
----------------------------------------
- nigeria_admin_codes.csv                             # list of filter codes
- nigeria_metrics_monthly_{MODEL}_fs{N}.csv           # per-month class-1 metrics
- nigeria_predictions_monthly_{MODEL}_fs{N}.csv       # Nigeria-filtered rows
- nigeria_results_summary.csv                         # one row per (model, fs)
- Nigeria_Model_Comparison_Table.xlsx                 # Table_Format-style workbook
"""

from __future__ import annotations

import csv
import os
from pathlib import Path

import geopandas as gpd
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


REPO_ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = REPO_ROOT / "Nigeria_experiments"
OUT_DIR.mkdir(parents=True, exist_ok=True)

NIGERIA_SHP = Path(
    r"C:\Users\swl00\IFPRI Dropbox\Weilun Shi\Google fund\Analysis"
    r"\1.Source Data\Outcome\FEWSNET_IPC"
    r"\FEWS NET Admin Boundaries\Nigeria.shp"
)

RESULTS_ROOT = (
    REPO_ROOT / "main_ablation_results" / "march2026_main_backup_month_ind_cont3"
)

MODEL_DIRS = {
    "GeoRF": "GF",
    "GeoXGB": "XGB",
    "GeoDT": "DT",
}

FS_VALUES = (1, 2, 3)
FS_TO_LAG = {1: 4, 2: 8, 3: 12}


def _confusion(y_true: pd.Series, y_pred: pd.Series) -> tuple[int, int, int, int]:
    tp = int(((y_true == 1) & (y_pred == 1)).sum())
    fp = int(((y_true == 0) & (y_pred == 1)).sum())
    fn = int(((y_true == 1) & (y_pred == 0)).sum())
    tn = int(((y_true == 0) & (y_pred == 0)).sum())
    return tp, fp, fn, tn


def _pr_f1(tp: int, fp: int, fn: int) -> tuple[float, float, float]:
    precision = tp / (tp + fp) if (tp + fp) > 0 else 0.0
    recall = tp / (tp + fn) if (tp + fn) > 0 else 0.0
    f1 = (
        2 * precision * recall / (precision + recall)
        if (precision + recall) > 0
        else 0.0
    )
    return precision, recall, f1


def _avg(vals: list[float]) -> float | None:
    vals = [v for v in vals if v is not None]
    return sum(vals) / len(vals) if vals else None


def _std(vals: list[float]) -> float:
    if len(vals) < 2:
        return 0.0
    m = sum(vals) / len(vals)
    return (sum((v - m) ** 2 for v in vals) / (len(vals) - 1)) ** 0.5


def load_nigeria_codes(shp_path: Path) -> set[int]:
    gdf = gpd.read_file(shp_path)
    if "admin_code" not in gdf.columns:
        raise KeyError(
            f"`admin_code` column not found in {shp_path}. "
            f"Available columns: {list(gdf.columns)}"
        )
    codes = {int(c) for c in gdf["admin_code"].dropna()}
    return codes


def compute_monthly_metrics(df_nig: pd.DataFrame) -> pd.DataFrame:
    """Replicate metrics_monthly.csv for Nigeria-only rows, for both models."""
    rows: list[dict] = []
    for month, g in df_nig.groupby("month_start", sort=True):
        y_true = g["y_true"].astype(int)
        for model_name, pred_col in [
            ("pooled", "y_pred_pooled"),
            ("partitioned", "y_pred_partitioned"),
        ]:
            y_pred = g[pred_col].astype(int)
            tp, fp, fn, tn = _confusion(y_true, y_pred)
            p, r, f1 = _pr_f1(tp, fp, fn)
            rows.append(
                {
                    "test_month": str(month)[:7],
                    "model": model_name,
                    "precision": p,
                    "recall": r,
                    "f1": f1,
                    "n": len(g),
                    "tp": tp,
                    "fp": fp,
                    "fn": fn,
                    "tn": tn,
                }
            )
    return pd.DataFrame(rows)


def aggregate_model_metrics(monthly: pd.DataFrame) -> dict:
    pooled = monthly[monthly["model"] == "pooled"]
    partitioned = monthly[monthly["model"] == "partitioned"]

    def col(sub: pd.DataFrame, c: str) -> list[float]:
        return [float(v) for v in sub[c].tolist()]

    return {
        "split_precision": _avg(col(partitioned, "precision")),
        "split_recall": _avg(col(partitioned, "recall")),
        "split_f1": _avg(col(partitioned, "f1")),
        "split_f1_std": _std(col(partitioned, "f1")),
        "pool_precision": _avg(col(pooled, "precision")),
        "pool_recall": _avg(col(pooled, "recall")),
        "pool_f1": _avg(col(pooled, "f1")),
        "pool_f1_std": _std(col(pooled, "f1")),
        "n_months": int(len(partitioned)),
    }


def process_model_fs(
    model_name: str,
    suffix: str,
    fs: int,
    nigeria_codes: set[int],
) -> dict | None:
    pred_path = (
        RESULTS_ROOT
        / f"result_partition_k40_compare_{suffix}_fs{fs}"
        / "predictions_monthly.csv"
    )
    if not pred_path.exists():
        print(f"  [skip] {model_name} fs{fs}: {pred_path} not found")
        return None

    df = pd.read_csv(pred_path)
    df_nig = df[df["FEWSNET_admin_code"].isin(nigeria_codes)].copy()

    pred_out = OUT_DIR / f"nigeria_predictions_monthly_{model_name}_fs{fs}.csv"
    df_nig.to_csv(pred_out, index=False)

    monthly = compute_monthly_metrics(df_nig)
    metrics_out = OUT_DIR / f"nigeria_metrics_monthly_{model_name}_fs{fs}.csv"
    monthly.to_csv(metrics_out, index=False)

    agg = aggregate_model_metrics(monthly)
    agg.update(
        {
            "model": model_name,
            "fs": fs,
            "lag_months": FS_TO_LAG[fs],
            "n_admin_units_matched": int(df_nig["FEWSNET_admin_code"].nunique()),
            "n_prediction_rows": int(len(df_nig)),
            "source_predictions_file": str(pred_path),
        }
    )

    print(
        f"  {model_name} fs{fs}: {agg['n_months']} months, "
        f"admin={agg['n_admin_units_matched']}, rows={agg['n_prediction_rows']}, "
        f"split_f1={agg['split_f1']:.4f}, pool_f1={agg['pool_f1']:.4f}"
    )
    return agg


def build_workbook(data: dict[tuple[str, int], dict]) -> Workbook:
    hfont = Font(bold=True, size=11)
    hfill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    mfont = Font(bold=True, size=11)
    thin = Border(
        left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin")
    )
    center = Alignment(horizontal="center")

    wb = Workbook()
    ws = wb.active
    ws.title = "Nigeria"

    ws.merge_cells("C1:E1")
    ws["C1"] = "Split Model"
    ws["C1"].font = hfont
    ws["C1"].alignment = center
    ws["C1"].fill = hfill

    ws.merge_cells("F1:H1")
    ws["F1"] = "Pooled Model (Non-split)"
    ws["F1"].font = hfont
    ws["F1"].alignment = center
    ws["F1"].fill = hfill

    headers = [
        "",
        "lag(months)",
        "Precision",
        "Recall",
        "F1",
        "precision",
        "recall",
        "F1",
        "F1 Improvement Percentage",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = center
        cell.border = thin

    row = 3
    for model in ["GeoRF", "GeoXGB", "GeoDT"]:
        first = True
        for fs in FS_VALUES:
            lag = FS_TO_LAG[fs]
            m = data.get((model, fs))
            if first:
                ws.cell(row=row, column=1, value=model).font = mfont
            ws.cell(row=row, column=2, value=lag).alignment = center

            if m:
                for ci, key in enumerate(
                    ["split_precision", "split_recall", "split_f1"], 3
                ):
                    v = m.get(key)
                    if v is not None:
                        ws.cell(row=row, column=ci, value=v)
                for ci, key in enumerate(
                    ["pool_precision", "pool_recall", "pool_f1"], 6
                ):
                    v = m.get(key)
                    if v is not None:
                        ws.cell(row=row, column=ci, value=v)
                sf, pf = m.get("split_f1"), m.get("pool_f1")
                if sf is not None and pf is not None and pf != 0:
                    ws.cell(row=row, column=9, value=(sf - pf) / pf)
                    ws.cell(row=row, column=9).number_format = "0.00%"

            for c in range(1, 10):
                ws.cell(row=row, column=c).border = thin
            first = False
            row += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    for col in "CDEFGH":
        ws.column_dimensions[col].width = 14
    ws.column_dimensions["I"].width = 26

    return wb


def main() -> None:
    print(f"Nigeria shapefile: {NIGERIA_SHP}")
    print(f"Results root    : {RESULTS_ROOT}")
    print(f"Output directory: {OUT_DIR}")
    print("-" * 80)

    nigeria_codes = load_nigeria_codes(NIGERIA_SHP)
    print(f"Loaded {len(nigeria_codes)} Nigeria admin_code values from shapefile")

    codes_csv = OUT_DIR / "nigeria_admin_codes.csv"
    with open(codes_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["FEWSNET_admin_code"])
        for code in sorted(nigeria_codes):
            w.writerow([code])
    print(f"Saved admin codes -> {codes_csv}")

    summary_rows: list[dict] = []
    data: dict[tuple[str, int], dict] = {}

    print("\nAggregating per (model, forecasting_scope):")
    for model_name, suffix in MODEL_DIRS.items():
        for fs in FS_VALUES:
            agg = process_model_fs(model_name, suffix, fs, nigeria_codes)
            if agg is not None:
                data[(model_name, fs)] = agg
                summary_rows.append(agg)

    if not summary_rows:
        print("\nNo results produced - nothing to write.")
        return

    summary_df = pd.DataFrame(summary_rows)
    preferred_cols = [
        "model",
        "fs",
        "lag_months",
        "n_months",
        "n_admin_units_matched",
        "n_prediction_rows",
        "split_precision",
        "split_recall",
        "split_f1",
        "split_f1_std",
        "pool_precision",
        "pool_recall",
        "pool_f1",
        "pool_f1_std",
        "source_predictions_file",
    ]
    ordered = [c for c in preferred_cols if c in summary_df.columns]
    summary_df = summary_df[ordered + [c for c in summary_df.columns if c not in ordered]]
    summary_path = OUT_DIR / "nigeria_results_summary.csv"
    summary_df.to_csv(summary_path, index=False)
    print(f"\nSaved per-model summary -> {summary_path}")

    wb = build_workbook(data)
    xlsx_path = OUT_DIR / "Nigeria_Model_Comparison_Table.xlsx"
    wb.save(xlsx_path)
    print(f"Saved comparison table -> {xlsx_path}")

    print("\n" + "=" * 130)
    fmt = (
        f"{'Model':<10} {'Lag':>3} | "
        f"{'P(split)':>10} {'R(split)':>10} {'F1(split)':>18} | "
        f"{'P(pool)':>10} {'R(pool)':>10} {'F1(pool)':>18} | {'Impr%':>8}"
    )
    print(fmt)
    print("-" * 130)
    for model in ["GeoRF", "GeoXGB", "GeoDT"]:
        for fs in FS_VALUES:
            m = data.get((model, fs))
            label = model if fs == 1 else ""
            lag = FS_TO_LAG[fs]
            if m:
                sp = f"{m['split_precision']:.4f}"
                sr = f"{m['split_recall']:.4f}"
                sf = f"{m['split_f1']:.4f}+-{m['split_f1_std']:.4f}"
                pp = f"{m['pool_precision']:.4f}"
                pr = f"{m['pool_recall']:.4f}"
                pf = f"{m['pool_f1']:.4f}+-{m['pool_f1_std']:.4f}"
                if m["split_f1"] is not None and m["pool_f1"]:
                    imp = f"{(m['split_f1'] - m['pool_f1']) / m['pool_f1'] * 100:+.2f}%"
                else:
                    imp = "   -"
                print(
                    f"{label:<10} {lag:>3} | {sp:>10} {sr:>10} {sf:>18} | "
                    f"{pp:>10} {pr:>10} {pf:>18} | {imp:>8}"
                )
            else:
                print(f"{label:<10} {lag:>3} | (no data)")
        print("-" * 130)


if __name__ == "__main__":
    main()
